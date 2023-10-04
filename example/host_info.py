# -*- coding: utf-8 -*-
import re
import sys
import os
import openpyxl

#IP地址分隔符和匹配规则
delimiter = ";"
pattern = r"^10\.\d{1,3}\.\d{1,3}\.\d{1,3}"

arguments = sys.argv
#current_directory = os.getcwd()
dir = os.getenv('RACK_EXECL_PATH', "/tmp")

if not os.path.exists(dir):
    # 如果文件夹不存在，使用创建文件夹
    os.makedirs(dir)
    print("创建文件夹%s" %(dir))

os.chdir(dir)
usage = """Usage:
    python %s arg1 $arg2 $arg3
  Description:
    arg1:                   第一个传参$1,指定主机ip所在列,默认是B列
    arg2:                   第二个传参$2,指定机架号的所在列,默认是G列
    arg3:                   第三个传参$3,指定主机机房所在列,默认是M列
Example:
    python %s B G M           指定ip,机架位,机房位置 所在的列
    python %s                 默认值执行
Env:
    RACK_EXECL_PATH           指定execl文件目录，默认/scripts""" %(arguments[0], arguments[0], arguments[0])

if len(arguments) > 1 and len(arguments) <= 3:
    if arguments[1] == "-h" or arguments[1] == "--help":
        print(usage)
        exit(0)
    # 将第一个参数赋值给 C
    ip_column = arguments[1]
    rack_column = arguments[2]
    idc_column = arguments[3]
else:
    ip_column = "B"
    rack_column = "G" 
    idc_column = "M"

# 打印Excel表中的所有表
workbook = openpyxl.load_workbook('rack-number.xlsx')

# 打印Excel表中的所有表
#print(workbook.sheetnames)

# 获取指定sheet表
sheet = workbook['Sheet1']

# 获取活动表
sheet = workbook.active
line_num = sheet.max_row
for row in range(line_num):
    cell_ip = sheet[ip_column + str(row + 1)].value
    cell_rack = sheet[rack_column +str(row + 1)].value
    cell_idc = sheet[idc_column +str(row + 1)].value
    parts = cell_ip.split(delimiter)
    for part in parts:
        ipmatches = re.findall(pattern, part)
        if ipmatches == "":
            pass
        else:
            for ip in ipmatches:
                idc = cell_idc.split('/')[0]
                format_metrics = 'resources_node_rack_number{host_ip="%s",cluster_idc="%s"} %s' %(ip, idc, cell_rack)
                #print(format_metrics.encode('utf-8'))
                print(format_metrics)
